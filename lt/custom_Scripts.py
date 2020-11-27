import openpyxl

def readExcel():
	file = r"C:\Projects\lti\static\resources\files\lead_Time_Indicator.xlsx"

	wb_Obj = openpyxl.load_workbook(file, data_only=True)

	sheet_Obj = wb_Obj.active

	c = sheet_Obj.max_column
	r = sheet_Obj.max_row

	data_Dict = {}
	for i in range(r):
		temp_List = []
		for j in range(c):
			cell_Obj = sheet_Obj.cell(row = i+1, column = j+1)
			temp_List.append(cell_Obj.value)
		ind = "r"+str(i)
		data_Dict[ind] = temp_List

	return data_Dict

def createDictFromExcel():
	file = r"C:\Projects\lti\static\resources\files\lead_Time_Indicator.xlsx"

	wb_Obj = openpyxl.load_workbook(file, data_only=True)

	sheet_Obj = wb_Obj['data bank']

	c = sheet_Obj.max_column
	r = sheet_Obj.max_row

	data_Dict = {}

	for i in range(r):
		for j in range(c):
			cell_Obj = sheet_Obj.cell(row = i+1, column = j+1)
			if(str(cell_Obj.fill.start_color.index) == "FFFFFF00"):
				# print("Cell Colour --> ", cell_Obj.fill.start_color.index)
				# print("Row --> ", i, "Column --> ", j)
				# print(cell_Obj.value)
				temp_Dict = {}
				for ind_i in range(i, r):
					c_Obj_1 = sheet_Obj.cell(row = ind_i + 2, column = j + 1)
					c_Obj_2 = sheet_Obj.cell(row = ind_i + 2, column = j + 2)
					# print(c_Obj_1.value, c_Obj_2.value)
					if(c_Obj_1.value == None or c_Obj_2.value == None):
						break
					else:
						temp_Dict[c_Obj_1.value] = c_Obj_2.value
				data_Dict[cell_Obj.value] = temp_Dict
	# print(data_Dict)
	return data_Dict

# createDictFromExcel()
